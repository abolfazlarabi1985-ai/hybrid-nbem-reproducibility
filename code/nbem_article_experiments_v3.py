"""
NBEM Article Reproducibility Experiments (v3)
================================================

A single, leakage-aware, resume-capable pipeline for the final NBEM-family article.
It intentionally excludes Random Forest, gradient boosting, XGBoost, LightGBM,
CatBoost, and tabular-foundation baselines because the final manuscript is framed
as an extension study within the Naive Bayes/NBEM family.

The pipeline produces:
- dataset audits, including the critical Diabetes and target-definition checks;
- five-seed results with weighted-F1 and macro-F1;
- WNB, NBEM, Adaptive Weighted NBEM, Hybrid NBEM, component models, and ablations;
- CV-sensitivity analyses for datasets that cannot support ordinary stratified 10-fold CV;
- article-ready CSV, XLSX, LaTeX tables, and Figures 1--3;
- statistical tests with Holm adjustment;
- execution-environment and reproducibility manifests.

Expected layout
---------------
project_root/
    datasets/processed/<dataset_folder>/*.csv
    code/nbem_article_experiments_v3.py
    config/dataset_config.json

Recommended final run
---------------------
python code/nbem_article_experiments_v3.py --project-root . --config config/dataset_config.json

The command is safe to resume. Use --no-resume only when intentionally replacing
all cached results.
"""

from __future__ import annotations

import argparse
import hashlib
import re
import shutil
import json
import math
import os
import platform
import sys
import time
import tracemalloc
import warnings
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

import numpy as np
import pandas as pd

import matplotlib.pyplot as plt

from sklearn.base import BaseEstimator
from sklearn.feature_selection import mutual_info_classif
from sklearn.impute import SimpleImputer
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import (
    accuracy_score,
    balanced_accuracy_score,
    f1_score,
    precision_score,
    recall_score,
    roc_auc_score,
)
from sklearn.model_selection import (
    GroupKFold,
    KFold,
    RepeatedStratifiedKFold,
    StratifiedKFold,
)
try:
    from sklearn.model_selection import StratifiedGroupKFold
except Exception:  # pragma: no cover
    StratifiedGroupKFold = None
from sklearn.naive_bayes import BernoulliNB, CategoricalNB, GaussianNB
from sklearn.neural_network import MLPClassifier
from sklearn.preprocessing import KBinsDiscretizer, LabelEncoder, OrdinalEncoder, StandardScaler

try:
    from scipy.stats import friedmanchisquare, rankdata, wilcoxon
except Exception:  # pragma: no cover
    friedmanchisquare = None
    rankdata = None
    wilcoxon = None


warnings.filterwarnings("ignore")

RANDOM_STATE = 42
EPS = 1e-12


# -----------------------------------------------------------------------------
# Canonical benchmark dataset list
# -----------------------------------------------------------------------------
# The original NBEM benchmark uses 20 datasets.  The script intentionally uses
# this whitelist by default, so root-level helper CSV files such as
# table_datasets_for_article.csv are never counted as datasets.
CANONICAL_20_DATASETS = [
    "abalone",
    "bank_marketing",
    "car_evaluation",
    "cardiotocography",
    "chronic_kidney_disease",
    "cirrhosis_patient_survival_prediction_dataset_1",
    "credit_approval",
    "diabetes_130_us_hospitals_for_years_1999_2008",
    "dry_bean_dataset",
    "ecoli",
    "glass_identification",
    "haberman_s_survival",
    "heart_disease",
    "heart_failure_clinical_records",
    "hepatitis",
    "ilpd_indian_liver_patient_dataset",
    "internet_advertisements",
    "mushroom",
    "predict_students_dropout_and_academic_success",
    "productivity_prediction_of_garment_employees",
]

# Some folders may have slightly longer names.  These prefixes make the script
# robust without silently adding extra datasets.
CANONICAL_DATASET_PREFIX_ALIASES = {
    "cirrhosis_patient_survival_prediction_dataset_1": ["cirrhosis_patient_survival_prediction"],
    "diabetes_130_us_hospitals_for_years_1999_2008": ["diabetes_130_us_hospitals"],
    "predict_students_dropout_and_academic_success": ["predict_students_dropout_and_academic"],
    "productivity_prediction_of_garment_employees": ["productivity_prediction_of_garment"],
}

ROOT_HELPER_CSV_PREFIXES = (
    "dataset_inventory",
    "table_datasets_for_article",
    "table_",
    "summary_",
)


# -----------------------------------------------------------------------------
# Utility functions
# -----------------------------------------------------------------------------


def set_reproducibility(seed: int = RANDOM_STATE) -> None:
    os.environ["PYTHONHASHSEED"] = str(seed)
    np.random.seed(seed)


def clean_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """Convert common missing markers and remove empty rows/columns."""
    missing_tokens = ["?", "NA", "N/A", "na", "null", "None", "", " "]
    return (
        df.replace(missing_tokens, np.nan)
        .dropna(axis=0, how="all")
        .dropna(axis=1, how="all")
    )


def infer_target(df: pd.DataFrame) -> str:
    """The benchmark files are assumed to store the target in the final column."""
    return str(df.columns[-1])


def detect_feature_groups(df: pd.DataFrame, target: str) -> Tuple[List[str], List[str], List[str]]:
    """Detect Boolean, categorical, and numerical features without using labels."""
    X = df.drop(columns=[target])
    bool_cols: List[str] = []
    cat_cols: List[str] = []
    num_cols: List[str] = []

    for col in X.columns:
        nunique = X[col].nunique(dropna=True)
        if nunique <= 2:
            bool_cols.append(col)
        elif pd.api.types.is_numeric_dtype(X[col]):
            num_cols.append(col)
        else:
            cat_cols.append(col)
    return bool_cols, cat_cols, num_cols


def stable_softmax(scores: np.ndarray, temperature: float = 1.0) -> np.ndarray:
    """Row-wise softmax with numerical stabilization."""
    scores = np.asarray(scores, dtype=float)
    temperature = max(float(temperature), EPS)
    z = scores / temperature
    z = z - np.nanmax(z, axis=1, keepdims=True)
    e = np.exp(z)
    denom = np.sum(e, axis=1, keepdims=True)
    return e / np.maximum(denom, EPS)


def probability_confidence_score(proba: np.ndarray, lambda_entropy: float = 0.5) -> np.ndarray:
    """
    Adaptive score g_t(x) used by Adaptive Weighted NBEM.

    For component t and sample x, let p_t(x) be the class-posterior vector. Define
    normalized entropy confidence C_t(x) and posterior margin M_t(x):

        C_t(x) = 1 - H(p_t(x)) / log(K)
        M_t(x) = p_(1)(x) - p_(2)(x)
        g_t(x) = lambda * C_t(x) + (1-lambda) * M_t(x)

    This function returns g_t(x) for all samples for one probabilistic component.
    """
    p = np.asarray(proba, dtype=float)
    p = np.clip(p, EPS, 1.0)
    p = p / np.maximum(p.sum(axis=1, keepdims=True), EPS)
    n_classes = p.shape[1]

    entropy = -np.sum(p * np.log(p), axis=1)
    if n_classes > 1:
        confidence = 1.0 - entropy / np.log(n_classes)
        sorted_p = np.sort(p, axis=1)
        margin = sorted_p[:, -1] - sorted_p[:, -2]
    else:
        confidence = np.ones(p.shape[0])
        margin = np.ones(p.shape[0])

    lam = float(lambda_entropy)
    lam = min(max(lam, 0.0), 1.0)
    score = lam * confidence + (1.0 - lam) * margin
    return np.nan_to_num(score, nan=0.0, posinf=1.0, neginf=0.0)


def align_proba(proba: np.ndarray, model_classes: Sequence[int], global_classes: Sequence[int]) -> np.ndarray:
    """Align a model probability matrix to a fixed global class order."""
    proba = np.asarray(proba, dtype=float)
    global_classes = np.asarray(global_classes)
    model_classes = np.asarray(model_classes)
    out = np.zeros((proba.shape[0], len(global_classes)), dtype=float)
    for j, cls in enumerate(model_classes):
        idx = np.where(global_classes == cls)[0]
        if len(idx):
            out[:, idx[0]] = proba[:, j]
    row_sum = out.sum(axis=1, keepdims=True)
    if np.any(row_sum <= EPS):
        out[row_sum[:, 0] <= EPS, :] = 1.0 / len(global_classes)
        row_sum = out.sum(axis=1, keepdims=True)
    return out / np.maximum(row_sum, EPS)


def safe_predict_proba(model: Any, X: Any, global_classes: Sequence[int]) -> np.ndarray:
    """Predict probabilities and align them to the global class set."""
    p = model.predict_proba(X)
    classes = getattr(model, "classes_", np.asarray(global_classes))
    return align_proba(p, classes, global_classes)


def safe_roc_auc(y_true: np.ndarray, proba: np.ndarray, global_classes: Sequence[int]) -> float:
    """ROC-AUC for binary and multi-class settings. Returns NaN if undefined."""
    try:
        present = np.unique(y_true)
        if len(present) < 2:
            return np.nan
        if len(global_classes) == 2:
            pos_idx = 1
            return float(roc_auc_score(y_true, proba[:, pos_idx]))
        return float(roc_auc_score(y_true, proba, multi_class="ovr", average="weighted", labels=list(global_classes)))
    except Exception:
        return np.nan


def compute_metrics(y_true: np.ndarray, y_pred: np.ndarray, proba: np.ndarray, global_classes: Sequence[int]) -> Dict[str, float]:
    """Return all article metrics with unambiguous weighted/macro names."""
    return {
        "accuracy": float(accuracy_score(y_true, y_pred)),
        "balanced_accuracy": float(balanced_accuracy_score(y_true, y_pred)),
        "precision_weighted": float(precision_score(y_true, y_pred, average="weighted", zero_division=0)),
        "recall_weighted": float(recall_score(y_true, y_pred, average="weighted", zero_division=0)),
        "f1_weighted": float(f1_score(y_true, y_pred, average="weighted", zero_division=0)),
        "precision_macro": float(precision_score(y_true, y_pred, average="macro", zero_division=0)),
        "recall_macro": float(recall_score(y_true, y_pred, average="macro", zero_division=0)),
        "f1_macro": float(f1_score(y_true, y_pred, average="macro", zero_division=0)),
        "roc_auc_weighted": safe_roc_auc(y_true, proba, global_classes),
    }


def time_and_memory_fit_predict(model: Any, fit_args: Tuple[Any, ...], pred_args: Tuple[Any, ...]) -> Tuple[Any, Any, Any, float, float, float]:
    """Return fitted model, pred, proba, training time, inference time, peak memory MB."""
    tracemalloc.start()
    t0 = time.perf_counter()
    model.fit(*fit_args)
    train_time = time.perf_counter() - t0

    t1 = time.perf_counter()
    pred = model.predict(*pred_args)
    proba = model.predict_proba(*pred_args)
    inference_time = time.perf_counter() - t1

    _, peak = tracemalloc.get_traced_memory()
    tracemalloc.stop()
    peak_memory_mb = peak / (1024.0 * 1024.0)
    return model, pred, proba, train_time, inference_time, peak_memory_mb


# -----------------------------------------------------------------------------
# Leakage-safe preprocessing
# -----------------------------------------------------------------------------


class NBEMPreprocessor:
    """
    Leakage-safe preprocessor fitted only on training folds.

    Outputs:
    - X_bool: encoded Boolean features for Bernoulli NB.
    - X_cat: encoded categorical features for Categorical NB.
    - X_num: standardized numerical features for Gaussian NB and MLP.
    - X_disc: discretized mixed representation for dependency-aware and categorical models.
    - X_cont: continuous mixed representation for MLP and general baselines.
    - X_tree: ordinal categorical/Boolean + raw imputed numerical features for boosting baselines.
    """

    def __init__(self, n_bins: int = 5):
        self.n_bins = int(n_bins)

    def fit(self, df: pd.DataFrame, target: str) -> "NBEMPreprocessor":
        self.target = target
        self.bool_cols, self.cat_cols, self.num_cols = detect_feature_groups(df, target)
        X = df.drop(columns=[target])

        self.bool_imp = SimpleImputer(strategy="most_frequent")
        self.cat_imp = SimpleImputer(strategy="most_frequent")
        self.num_imp = SimpleImputer(strategy="median")

        self.bool_enc = OrdinalEncoder(handle_unknown="use_encoded_value", unknown_value=-1)
        self.cat_enc = OrdinalEncoder(handle_unknown="use_encoded_value", unknown_value=-1)
        self.scaler = StandardScaler()
        self.binner = KBinsDiscretizer(n_bins=self.n_bins, encode="ordinal", strategy="quantile")

        if self.bool_cols:
            B = self.bool_imp.fit_transform(X[self.bool_cols].astype(str))
            self.bool_enc.fit(B)
        if self.cat_cols:
            C = self.cat_imp.fit_transform(X[self.cat_cols].astype(str))
            self.cat_enc.fit(C)
        if self.num_cols:
            N = self.num_imp.fit_transform(X[self.num_cols])
            self.scaler.fit(N)
            # Quantile binning can fail when features are nearly constant. Keep it safe.
            try:
                self.binner.fit(N)
                self._binner_is_fitted = True
            except Exception:
                self._binner_is_fitted = False
        else:
            self._binner_is_fitted = False
        return self

    def transform(self, df: pd.DataFrame) -> Dict[str, Any]:
        if not hasattr(self, "target"):
            raise RuntimeError("NBEMPreprocessor must be fitted before transform.")
        X = df.drop(columns=[self.target])
        n = len(df)

        X_bool = np.empty((n, 0), dtype=int)
        X_cat = np.empty((n, 0), dtype=int)
        X_num = np.empty((n, 0), dtype=float)
        X_num_raw = np.empty((n, 0), dtype=float)
        X_num_disc = np.empty((n, 0), dtype=int)

        cont_parts: List[np.ndarray] = []
        disc_parts: List[np.ndarray] = []
        tree_parts: List[np.ndarray] = []

        if self.bool_cols:
            B = self.bool_imp.transform(X[self.bool_cols].astype(str))
            B = self.bool_enc.transform(B)
            B = np.maximum(B, 0).astype(int)
            X_bool = B
            cont_parts.append(B.astype(float))
            disc_parts.append(B)
            tree_parts.append(B.astype(float))

        if self.cat_cols:
            C = self.cat_imp.transform(X[self.cat_cols].astype(str))
            C = self.cat_enc.transform(C)
            C = np.maximum(C, 0).astype(int)
            X_cat = C
            cont_parts.append(C.astype(float))
            disc_parts.append(C)
            tree_parts.append(C.astype(float))

        if self.num_cols:
            N = self.num_imp.transform(X[self.num_cols])
            X_num_raw = N.astype(float)
            X_num = self.scaler.transform(N)
            if self._binner_is_fitted:
                X_num_disc = self.binner.transform(N).astype(int)
            else:
                X_num_disc = np.zeros_like(N, dtype=int)
            cont_parts.append(X_num)
            disc_parts.append(X_num_disc)
            tree_parts.append(X_num_raw)

        X_cont = np.hstack(cont_parts) if cont_parts else np.zeros((n, 1), dtype=float)
        X_disc = np.hstack(disc_parts) if disc_parts else np.zeros((n, 1), dtype=int)
        X_tree = np.hstack(tree_parts) if tree_parts else np.zeros((n, 1), dtype=float)

        return {
            "X_cont": X_cont.astype(float),
            "X_tree": X_tree.astype(float),
            "X_disc": X_disc.astype(int),
            "X_bool": X_bool.astype(int),
            "X_cat": X_cat.astype(int),
            "X_num": X_num.astype(float),
            "X_num_raw": X_num_raw.astype(float),
            "bool_cols": self.bool_cols,
            "cat_cols": self.cat_cols,
            "num_cols": self.num_cols,
        }


# -----------------------------------------------------------------------------
# NBEM family models
# -----------------------------------------------------------------------------


class WeightedCategoricalNB:
    """Categorical NB with mutual-information feature weights."""

    def __init__(self, alpha: float = 1.0, random_state: int = RANDOM_STATE):
        self.alpha = float(alpha)
        self.random_state = int(random_state)

    def fit(self, X: np.ndarray, y: np.ndarray) -> "WeightedCategoricalNB":
        X = np.asarray(X).astype(int)
        y = np.asarray(y)
        self.classes_, y_idx = np.unique(y, return_inverse=True)
        self.n_classes_ = len(self.classes_)
        self.n_features_ = X.shape[1]

        try:
            w = mutual_info_classif(X, y, discrete_features=True, random_state=self.random_state)
            w = np.nan_to_num(w, nan=0.0, posinf=0.0, neginf=0.0)
            if np.sum(w) <= EPS:
                w = np.ones(self.n_features_)
            self.feature_weights_ = w / np.sum(w)
        except Exception:
            self.feature_weights_ = np.ones(self.n_features_) / max(self.n_features_, 1)

        class_counts = np.bincount(y_idx, minlength=self.n_classes_)
        self.class_log_prior_ = np.log((class_counts + self.alpha) / (len(y) + self.alpha * self.n_classes_))

        self.feature_log_prob_: List[np.ndarray] = []
        for j in range(self.n_features_):
            max_val = int(np.nanmax(X[:, j])) if X.shape[0] else 0
            n_cat = max(max_val + 1, 1)
            probs = np.zeros((self.n_classes_, n_cat), dtype=float)
            for c in range(self.n_classes_):
                vals = X[y_idx == c, j]
                counts = np.bincount(vals, minlength=n_cat)
                probs[c] = np.log((counts + self.alpha) / (counts.sum() + self.alpha * n_cat))
            self.feature_log_prob_.append(probs)
        return self

    def predict_proba(self, X: np.ndarray) -> np.ndarray:
        X = np.asarray(X).astype(int)
        logp = np.tile(self.class_log_prior_, (X.shape[0], 1))
        for j in range(self.n_features_):
            probs = self.feature_log_prob_[j]
            vals = np.clip(X[:, j], 0, probs.shape[1] - 1)
            for c in range(self.n_classes_):
                logp[:, c] += self.feature_weights_[j] * probs[c, vals]
        logp -= np.max(logp, axis=1, keepdims=True)
        p = np.exp(logp)
        return p / np.maximum(p.sum(axis=1, keepdims=True), EPS)

    def predict(self, X: np.ndarray) -> np.ndarray:
        return self.classes_[np.argmax(self.predict_proba(X), axis=1)]


class NBEMClassifier:
    """Reference NBEM: Boolean, categorical, and numerical NB components with feature-proportion weights."""

    def fit(self, data: Dict[str, Any], y: np.ndarray) -> "NBEMClassifier":
        self.classes_ = np.unique(y)
        self.models_: Dict[str, Any] = {}
        sizes = {
            "bool": data["X_bool"].shape[1],
            "cat": data["X_cat"].shape[1],
            "num": data["X_num"].shape[1],
        }
        total = sum(sizes.values()) or 1
        self.component_weights_ = {k: v / total for k, v in sizes.items()}

        if sizes["bool"] > 0:
            self.models_["bool"] = BernoulliNB().fit(data["X_bool"], y)
        if sizes["cat"] > 0:
            self.models_["cat"] = CategoricalNB().fit(data["X_cat"], y)
        if sizes["num"] > 0:
            self.models_["num"] = GaussianNB().fit(data["X_num"], y)
        return self

    def _component_probas(self, data: Dict[str, Any]) -> Dict[str, np.ndarray]:
        probas: Dict[str, np.ndarray] = {}
        if "bool" in self.models_:
            probas["bool"] = safe_predict_proba(self.models_["bool"], data["X_bool"], self.classes_)
        if "cat" in self.models_:
            probas["cat"] = safe_predict_proba(self.models_["cat"], data["X_cat"], self.classes_)
        if "num" in self.models_:
            probas["num"] = safe_predict_proba(self.models_["num"], data["X_num"], self.classes_)
        return probas

    def predict_proba(self, data: Dict[str, Any]) -> np.ndarray:
        probas = self._component_probas(data)
        if not probas:
            return np.ones((data["X_cont"].shape[0], len(self.classes_))) / len(self.classes_)
        weights = np.asarray([self.component_weights_[k] for k in probas.keys()], dtype=float)
        weights = weights / np.maximum(weights.sum(), EPS)
        fused = np.zeros_like(next(iter(probas.values())))
        for w, p in zip(weights, probas.values()):
            fused += w * p
        return fused / np.maximum(fused.sum(axis=1, keepdims=True), EPS)

    def predict(self, data: Dict[str, Any]) -> np.ndarray:
        return self.classes_[np.argmax(self.predict_proba(data), axis=1)]


class AdaptiveWeightedNBEM:
    """
    Adaptive Weighted NBEM using the professor-requested exact adaptive function.

    For each sample x and component t, the model computes posterior p_t(x), confidence score g_t(x),
    and adaptive weight alpha_t(x)=softmax(g_t(x)/tau). The final posterior is a convex mixture
    of component posteriors with sample-specific weights.
    """

    def __init__(self, temperature: float = 1.0, lambda_entropy: float = 0.5):
        self.temperature = float(temperature)
        self.lambda_entropy = float(lambda_entropy)

    def fit(self, data: Dict[str, Any], y: np.ndarray) -> "AdaptiveWeightedNBEM":
        self.classes_ = np.unique(y)
        self.models_: Dict[str, Any] = {}
        if data["X_bool"].shape[1] > 0:
            self.models_["bool"] = BernoulliNB().fit(data["X_bool"], y)
        if data["X_cat"].shape[1] > 0:
            self.models_["cat"] = CategoricalNB().fit(data["X_cat"], y)
        if data["X_num"].shape[1] > 0:
            self.models_["num"] = GaussianNB().fit(data["X_num"], y)
        return self

    def _component_probas(self, data: Dict[str, Any]) -> Dict[str, np.ndarray]:
        probas: Dict[str, np.ndarray] = {}
        if "bool" in self.models_:
            probas["bool"] = safe_predict_proba(self.models_["bool"], data["X_bool"], self.classes_)
        if "cat" in self.models_:
            probas["cat"] = safe_predict_proba(self.models_["cat"], data["X_cat"], self.classes_)
        if "num" in self.models_:
            probas["num"] = safe_predict_proba(self.models_["num"], data["X_num"], self.classes_)
        return probas

    def adaptive_weights(self, data: Dict[str, Any]) -> Tuple[List[str], np.ndarray, Dict[str, np.ndarray]]:
        probas = self._component_probas(data)
        names = list(probas.keys())
        if not names:
            n = data["X_cont"].shape[0]
            return [], np.ones((n, 1)), probas
        scores = np.vstack([
            probability_confidence_score(probas[name], lambda_entropy=self.lambda_entropy)
            for name in names
        ]).T
        weights = stable_softmax(scores, temperature=self.temperature)
        return names, weights, probas

    def predict_proba(self, data: Dict[str, Any]) -> np.ndarray:
        names, weights, probas = self.adaptive_weights(data)
        if not names:
            return np.ones((data["X_cont"].shape[0], len(self.classes_))) / len(self.classes_)
        fused = np.zeros_like(probas[names[0]])
        for j, name in enumerate(names):
            fused += weights[:, [j]] * probas[name]
        return fused / np.maximum(fused.sum(axis=1, keepdims=True), EPS)

    def predict(self, data: Dict[str, Any]) -> np.ndarray:
        return self.classes_[np.argmax(self.predict_proba(data), axis=1)]


class DependencyAwareNBEM:
    """
    Dependency-aware categorical NB based on selected pairwise interaction features.

    Important implementation detail for high-dimensional datasets
    ------------------------------------------------------------
    The Internet Advertisements dataset has more than 1500 features.  Computing all
    pairwise correlations would require more than 1.2 million feature pairs per fold
    and becomes impractical in a full 10-fold experiment.  To keep the method faithful
    and scalable, candidate features are first selected on the training fold using
    mutual information with the class label, and interactions are then selected only
    among these candidates.  This is leakage-safe because candidate selection is done
    only inside the training fold.
    """

    def __init__(
        self,
        max_interactions: int = 10,
        max_candidate_features: int = 80,
        alpha: float = 1.0,
        random_state: int = RANDOM_STATE,
    ):
        self.max_interactions = int(max_interactions)
        self.max_candidate_features = int(max_candidate_features)
        self.alpha = float(alpha)
        self.random_state = int(random_state)
        self.base_ = WeightedCategoricalNB(alpha=alpha, random_state=random_state)

    def _select_candidate_features(self, X: np.ndarray, y: np.ndarray) -> np.ndarray:
        X = np.asarray(X).astype(int)
        m = X.shape[1]
        if m <= self.max_candidate_features:
            return np.arange(m, dtype=int)

        non_constant = np.array([np.nanstd(X[:, j]) > EPS for j in range(m)])
        valid_idx = np.where(non_constant)[0]
        if len(valid_idx) <= self.max_candidate_features:
            return valid_idx.astype(int)

        try:
            mi = mutual_info_classif(
                X[:, valid_idx],
                y,
                discrete_features=True,
                random_state=self.random_state,
            )
            mi = np.nan_to_num(mi, nan=0.0, posinf=0.0, neginf=0.0)
            order = np.argsort(mi)[::-1]
            selected = valid_idx[order[: self.max_candidate_features]]
        except Exception:
            # Robust fallback: select features with the largest variance.
            var = np.nanvar(X[:, valid_idx], axis=0)
            order = np.argsort(var)[::-1]
            selected = valid_idx[order[: self.max_candidate_features]]
        return np.asarray(selected, dtype=int)

    def _fit_pairs(self, X: np.ndarray, y: np.ndarray) -> None:
        X = np.asarray(X).astype(int)
        self.selected_features_ = self._select_candidate_features(X, y)
        if X.shape[1] < 2 or len(self.selected_features_) < 2 or self.max_interactions <= 0:
            self.pairs_ = []
            self.cardinalities_ = {}
            return

        Xc = X[:, self.selected_features_].astype(float)
        # Standardize only for correlation scoring; this does not alter the actual features.
        std = Xc.std(axis=0)
        keep = std > EPS
        selected = self.selected_features_[keep]
        Xc = Xc[:, keep]
        if Xc.shape[1] < 2:
            self.pairs_ = []
            self.cardinalities_ = {}
            return

        try:
            corr = np.corrcoef(Xc, rowvar=False)
            corr = np.nan_to_num(np.abs(corr), nan=0.0, posinf=0.0, neginf=0.0)
        except Exception:
            corr = np.zeros((Xc.shape[1], Xc.shape[1]), dtype=float)

        iu = np.triu_indices_from(corr, k=1)
        scores = corr[iu]
        if scores.size == 0:
            self.pairs_ = []
            self.cardinalities_ = {}
            return

        k = min(self.max_interactions, scores.size)
        # Efficient top-k without sorting millions of pairs.
        top_local = np.argpartition(scores, -k)[-k:]
        top_local = top_local[np.argsort(scores[top_local])[::-1]]
        self.pairs_ = [(int(selected[iu[0][idx]]), int(selected[iu[1][idx]])) for idx in top_local]
        self.cardinalities_ = {j: int(np.nanmax(X[:, j])) + 1 for _, j in self.pairs_}

    def _make_interactions(self, X: np.ndarray, fit: bool = False, y: Optional[np.ndarray] = None) -> np.ndarray:
        X = np.asarray(X).astype(int)
        if fit:
            if y is None:
                raise ValueError("y is required when fitting dependency-aware interactions.")
            self._fit_pairs(X, y)

        inter: List[np.ndarray] = []
        for i, j in getattr(self, "pairs_", []):
            base = self.cardinalities_.get(j, int(np.nanmax(X[:, j])) + 1)
            z = X[:, i] * max(base, 1) + X[:, j]
            inter.append(z.reshape(-1, 1))
        return np.hstack([X] + inter) if inter else X

    def fit(self, X: np.ndarray, y: np.ndarray) -> "DependencyAwareNBEM":
        X2 = self._make_interactions(X, fit=True, y=y)
        self.base_.fit(X2, y)
        self.classes_ = self.base_.classes_
        return self

    def predict_proba(self, X: np.ndarray) -> np.ndarray:
        return self.base_.predict_proba(self._make_interactions(X, fit=False))

    def predict(self, X: np.ndarray) -> np.ndarray:
        return self.classes_[np.argmax(self.predict_proba(X), axis=1)]


class ProbabilisticDeepNBEM:
    """MLP probabilistic component used in Hybrid NBEM."""

    def __init__(self, hidden_layer_sizes: Tuple[int, int] = (128, 64), max_iter: int = 300, random_state: int = RANDOM_STATE):
        self.hidden_layer_sizes = hidden_layer_sizes
        self.max_iter = int(max_iter)
        self.random_state = int(random_state)

    def fit(self, X: np.ndarray, y: np.ndarray) -> "ProbabilisticDeepNBEM":
        self.model_ = MLPClassifier(
            hidden_layer_sizes=self.hidden_layer_sizes,
            activation="relu",
            solver="adam",
            alpha=1e-4,
            batch_size="auto",
            learning_rate="adaptive",
            learning_rate_init=1e-3,
            max_iter=self.max_iter,
            early_stopping=True,
            validation_fraction=0.1,
            n_iter_no_change=20,
            random_state=self.random_state,
        )
        self.model_.fit(X, y)
        self.classes_ = self.model_.classes_
        return self

    def predict_proba(self, X: np.ndarray) -> np.ndarray:
        return self.model_.predict_proba(X)

    def predict(self, X: np.ndarray) -> np.ndarray:
        return self.model_.predict(X)


class HybridNBEM:
    """
    Hybrid NBEM with configurable components for full model and ablation variants.

    Components:
    - dependency-aware NBEM
    - adaptive weighted NBEM
    - probabilistic deep NBEM
    Meta-fusion:
    - Logistic Regression over concatenated probability vectors
    """

    def __init__(
        self,
        use_dependency: bool = True,
        use_adaptive: bool = True,
        use_deep: bool = True,
        max_interactions: int = 10,
        temperature: float = 1.0,
        lambda_entropy: float = 0.5,
        mlp_hidden: Tuple[int, int] = (128, 64),
        mlp_max_iter: int = 300,
        random_state: int = RANDOM_STATE,
    ):
        self.use_dependency = bool(use_dependency)
        self.use_adaptive = bool(use_adaptive)
        self.use_deep = bool(use_deep)
        self.max_interactions = int(max_interactions)
        self.temperature = float(temperature)
        self.lambda_entropy = float(lambda_entropy)
        self.mlp_hidden = mlp_hidden
        self.mlp_max_iter = int(mlp_max_iter)
        self.random_state = int(random_state)

    def fit(self, data: Dict[str, Any], y: np.ndarray) -> "HybridNBEM":
        self.global_classes_ = np.unique(y)
        self.components_: List[Tuple[str, Any, str]] = []

        if self.use_dependency:
            dep = DependencyAwareNBEM(max_interactions=self.max_interactions, random_state=self.random_state).fit(data["X_disc"], y)
            self.components_.append(("Dependency", dep, "X_disc"))

        if self.use_adaptive:
            aw = AdaptiveWeightedNBEM(temperature=self.temperature, lambda_entropy=self.lambda_entropy).fit(data, y)
            self.components_.append(("Adaptive", aw, "data"))

        if self.use_deep:
            deep = ProbabilisticDeepNBEM(hidden_layer_sizes=self.mlp_hidden, max_iter=self.mlp_max_iter, random_state=self.random_state).fit(data["X_cont"], y)
            self.components_.append(("Deep", deep, "X_cont"))

        if not self.components_:
            # Safe fallback: if no component is selected, use NBEM.
            nbem = NBEMClassifier().fit(data, y)
            self.components_.append(("NBEM", nbem, "data"))

        P = self._stack_component_probas(data)
        self.meta_ = LogisticRegression(max_iter=1000, class_weight=None, random_state=self.random_state).fit(P, y)
        self.classes_ = self.meta_.classes_
        return self

    def _component_predict_proba(self, component: Any, mode: str, data: Dict[str, Any]) -> np.ndarray:
        if mode == "data":
            p = component.predict_proba(data)
            cls = getattr(component, "classes_", self.global_classes_)
        else:
            p = component.predict_proba(data[mode])
            cls = getattr(component, "classes_", self.global_classes_)
        return align_proba(p, cls, self.global_classes_)

    def _stack_component_probas(self, data: Dict[str, Any]) -> np.ndarray:
        return np.hstack([self._component_predict_proba(comp, mode, data) for _, comp, mode in self.components_])

    def predict_proba(self, data: Dict[str, Any]) -> np.ndarray:
        P = self._stack_component_probas(data)
        return safe_predict_proba(self.meta_, P, self.global_classes_)

    def predict(self, data: Dict[str, Any]) -> np.ndarray:
        P = self._stack_component_probas(data)
        return self.meta_.predict(P)


# -----------------------------------------------------------------------------
# Final article experiment configuration (NB/NBEM family only)
# -----------------------------------------------------------------------------

ARTICLE_MODELS = [
    "Gaussian NB",
    "Bernoulli NB",
    "Categorical NB",
    "WNB",
    "NBEM",
    "Dependency-Aware NBEM",
    "Adaptive Weighted NBEM",
    "Probabilistic Deep NBEM",
    "Hybrid w/o Dependency",
    "Hybrid w/o Adaptive",
    "Hybrid w/o Deep",
    "Hybrid NBEM",
]

KEY_MODELS = ["WNB", "NBEM", "Adaptive Weighted NBEM", "Hybrid NBEM"]
ABLATION_MODELS = [
    "Hybrid NBEM",
    "Hybrid w/o Dependency",
    "Hybrid w/o Deep",
    "Hybrid w/o Adaptive",
    "NBEM",
    "Adaptive Weighted NBEM",
    "Probabilistic Deep NBEM",
    "Dependency-Aware NBEM",
]

METRIC_COLUMNS = [
    "accuracy",
    "balanced_accuracy",
    "precision_weighted",
    "recall_weighted",
    "f1_weighted",
    "precision_macro",
    "recall_macro",
    "f1_macro",
    "roc_auc_weighted",
    "train_time_s",
    "inference_time_s",
    "peak_memory_mb",
]

DISPLAY_DATASET_NAMES = {
    "diabetes_130_us_hospitals_for_years_1999_2008": "Diabetes",
    "cardiotocography": "Fetal Health",
    "heart_disease": "Heart Disease",
    "predict_students_dropout_and_academic_success": "Students",
    "bank_marketing": "Bank",
    "heart_failure_clinical_records": "Heart Failure",
    "internet_advertisements": "Internet Ads",
    "chronic_kidney_disease": "Kidney Disease",
    "mushroom": "Mushroom",
    "dry_bean_dataset": "Dry Beans",
    "abalone": "Abalone",
    "car_evaluation": "Car Evaluation",
    "credit_approval": "Credit Approval",
    "ecoli": "E. coli",
    "glass_identification": "Glass Identification",
    "hepatitis": "Hepatitis",
    "productivity_prediction_of_garment_employees": "Productivity Prediction",
    "haberman_s_survival": "Haberman Survival",
    "ilpd_indian_liver_patient_dataset": "Indian Liver Patients",
    "cirrhosis_patient_survival_prediction_dataset_1": "Cirrhosis Survival",
}


@dataclass
class PreparedDataset:
    key: str
    display_name: str
    path: Path
    df: pd.DataFrame
    target: str
    groups: Optional[np.ndarray]
    group_column: Optional[str]
    dropped_columns: List[str]
    target_transform: str
    audit: Dict[str, Any]
    leakage_rows: List[Dict[str, Any]]
    critical_issues: List[str]


@dataclass
class CVPlan:
    protocol: str
    n_splits: int
    splits: List[Tuple[np.ndarray, np.ndarray]]
    stratified: bool
    grouped: bool


def normalize_name(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value).strip().lower())


def resolve_column(columns: Sequence[Any], candidates: Sequence[str]) -> Optional[str]:
    by_norm = {normalize_name(c): str(c) for c in columns}
    for candidate in candidates:
        found = by_norm.get(normalize_name(candidate))
        if found is not None:
            return found
    return None


def _canonical_key(folder_name: str) -> Optional[str]:
    name = folder_name.strip().lower()
    for key in CANONICAL_20_DATASETS:
        if name == key.lower():
            return key
    for key, prefixes in CANONICAL_DATASET_PREFIX_ALIASES.items():
        for prefix in prefixes:
            if name.startswith(prefix.lower()):
                return key
    return None


def _choose_dataset_csv(folder: Path) -> Optional[Path]:
    files = [p for p in folder.glob("*.csv") if not p.name.lower().startswith(ROOT_HELPER_CSV_PREFIXES)]
    if not files:
        return None
    normalized_folder = folder.name.lower().replace("-", "_")
    for file in files:
        stem = file.stem.lower().replace("-", "_")
        if stem == normalized_folder or normalized_folder.startswith(stem) or stem.startswith(normalized_folder[:20]):
            return file
    return max(files, key=lambda x: x.stat().st_size)


def find_datasets(project_root: Path, use_canonical_20: bool = True) -> List[Path]:
    data_dir = project_root / "datasets" / "processed"
    if not data_dir.exists():
        return []
    folders = sorted([p for p in data_dir.iterdir() if p.is_dir()])
    if not use_canonical_20:
        return [f for folder in folders if (f := _choose_dataset_csv(folder)) is not None]
    folder_by_key: Dict[str, Path] = {}
    for folder in folders:
        key = _canonical_key(folder.name)
        if key is not None and key not in folder_by_key:
            folder_by_key[key] = folder
    files: List[Path] = []
    missing: List[str] = []
    for key in CANONICAL_20_DATASETS:
        folder = folder_by_key.get(key)
        file = _choose_dataset_csv(folder) if folder is not None else None
        if file is None:
            missing.append(key)
        else:
            files.append(file)
    if missing:
        print("WARNING: canonical datasets not found: " + ", ".join(missing))
    return files


def canonical_dataset_key(path: Path) -> str:
    folder = path.parent.name if path.parent.name != "processed" else path.stem
    return _canonical_key(folder) or folder


def load_json(path: Path) -> Dict[str, Any]:
    if not path.exists():
        raise FileNotFoundError(f"Configuration file not found: {path}")
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def parse_bins(values: Sequence[Any]) -> List[float]:
    parsed: List[float] = []
    for value in values:
        if isinstance(value, str) and value.lower() in {"-inf", "-infinity"}:
            parsed.append(float("-inf"))
        elif isinstance(value, str) and value.lower() in {"inf", "+inf", "infinity", "+infinity"}:
            parsed.append(float("inf"))
        else:
            parsed.append(float(value))
    return parsed


def apply_target_transform(series: pd.Series, specification: Optional[Dict[str, Any]]) -> Tuple[pd.Series, str]:
    if not specification:
        return series, "none"
    kind = str(specification.get("type", "none")).lower()
    if kind == "none":
        return series, "none"
    if kind == "fixed_bins":
        bins = parse_bins(specification["bins"])
        labels = specification.get("labels")
        numeric = pd.to_numeric(series, errors="coerce")
        transformed = pd.cut(numeric, bins=bins, labels=labels, include_lowest=True, right=False)
        desc = f"fixed_bins(bins={bins}, labels={labels}, right=False)"
        return transformed.astype(object), desc
    if kind == "mapping":
        mapping = {str(k): str(v) for k, v in specification.get("mapping", {}).items()}
        transformed = series.astype(str).map(mapping)
        return transformed, f"mapping({len(mapping)} entries)"
    raise ValueError(f"Unsupported target transform: {kind}")


def feature_target_purity(feature: pd.Series, target: pd.Series) -> float:
    tmp = pd.DataFrame({"x": feature.astype(str), "y": target.astype(str)}).dropna()
    if tmp.empty:
        return np.nan
    counts = tmp.groupby(["x", "y"], dropna=False).size().unstack(fill_value=0)
    return float(counts.max(axis=1).sum() / counts.to_numpy().sum())


def audit_feature_leakage(df: pd.DataFrame, target: str, max_rows: int = 20000) -> List[Dict[str, Any]]:
    work = df if len(df) <= max_rows else df.sample(max_rows, random_state=RANDOM_STATE)
    y = work[target]
    rows: List[Dict[str, Any]] = []
    for col in work.columns:
        if col == target:
            continue
        x = work[col]
        nunique = int(x.nunique(dropna=True))
        unique_ratio = float(nunique / max(len(work), 1))
        exact_match = float((x.astype(str) == y.astype(str)).mean())
        purity = np.nan
        if 1 < nunique <= 500:
            try:
                purity = feature_target_purity(x, y)
            except Exception:
                purity = np.nan
        name_norm = normalize_name(col)
        id_like = bool(
            name_norm.endswith("id")
            or "identifier" in name_norm
            or name_norm in {"patientnbr", "encounterid", "recordid", "index"}
        )
        flagged = bool(exact_match >= 0.995 or (np.isfinite(purity) and purity >= 0.995) or (id_like and unique_ratio > 0.5))
        if flagged:
            rows.append({
                "feature": str(col),
                "n_unique": nunique,
                "unique_ratio": unique_ratio,
                "exact_target_match": exact_match,
                "target_determinism_purity": purity,
                "id_like": id_like,
                "flag_reason": "; ".join([
                    r for r, ok in [
                        ("exact/near-exact target copy", exact_match >= 0.995),
                        ("near-deterministic mapping to target", np.isfinite(purity) and purity >= 0.995),
                        ("high-cardinality identifier-like feature", id_like and unique_ratio > 0.5),
                    ] if ok
                ]),
            })
    return rows


def get_dataset_override(config: Dict[str, Any], key: str, path: Path) -> Dict[str, Any]:
    overrides = config.get("dataset_overrides", {})
    if key in overrides:
        return overrides[key]
    folder_norm = normalize_name(path.parent.name)
    for name, value in overrides.items():
        if normalize_name(name) == folder_norm:
            return value
    return {}


def prepare_dataset(path: Path, config: Dict[str, Any]) -> PreparedDataset:
    key = canonical_dataset_key(path)
    override = get_dataset_override(config, key, path)
    raw = clean_dataframe(pd.read_csv(path))
    if raw.shape[1] < 2:
        raise ValueError(f"Dataset {path} has fewer than two columns.")

    inferred_target = infer_target(raw)
    target_candidates = override.get("target_candidates", [])
    target = resolve_column(raw.columns, target_candidates) if target_candidates else inferred_target
    critical: List[str] = []
    if target is None:
        critical.append(f"Configured target candidates were not found: {target_candidates}")
        target = inferred_target

    forbidden = {normalize_name(x) for x in override.get("forbidden_targets", [])}
    if normalize_name(target) in forbidden:
        critical.append(f"Forbidden/questionable target selected: {target}")

    group_column = resolve_column(raw.columns, override.get("group_column_candidates", []))
    groups = raw[group_column].astype(str).to_numpy() if group_column is not None else None

    transformed_target, transform_desc = apply_target_transform(raw[target], override.get("target_transform"))
    raw[target] = transformed_target
    before_drop = len(raw)
    raw = raw.dropna(subset=[target]).reset_index(drop=True)
    if groups is not None:
        valid_mask = pd.notna(transformed_target).to_numpy()
        groups = groups[valid_mask]

    drop_candidates = list(override.get("drop_feature_candidates", []))
    # Alternative labels must not be used as predictors.
    drop_candidates += list(override.get("alternative_target_candidates", []))
    dropped: List[str] = []
    for candidate in drop_candidates:
        col = resolve_column(raw.columns, [candidate])
        if col is not None and col != target:
            raw = raw.drop(columns=[col])
            dropped.append(col)

    n_rows = int(len(raw))
    n_features = int(raw.shape[1] - 1)
    n_unique = int(raw[target].nunique(dropna=True))
    counts = raw[target].astype(str).value_counts(dropna=False)
    min_class = int(counts.min()) if len(counts) else 0
    max_class = int(counts.max()) if len(counts) else 0
    imbalance_ratio = float(max_class / max(min_class, 1)) if len(counts) else np.nan

    numeric_target = pd.to_numeric(raw[target], errors="coerce")
    numeric_fraction = float(numeric_target.notna().mean())
    unique_ratio = float(n_unique / max(n_rows, 1))
    if n_unique < 2:
        critical.append("Target contains fewer than two classes.")
    if n_unique > max(50, int(math.sqrt(max(n_rows, 1))) + 5):
        critical.append(
            f"Target has {n_unique} distinct values for {n_rows} rows; this looks continuous or incorrectly defined for classification."
        )
    if numeric_fraction > 0.95 and unique_ratio > 0.20 and not override.get("target_transform"):
        critical.append("Target appears continuous and no classification transform is configured.")

    leakage_rows = audit_feature_leakage(raw, target)
    target_norm = normalize_name(target)
    severe_leakage = []
    for row in leakage_rows:
        feature_norm = normalize_name(row["feature"])
        target_named = target_norm and (target_norm in feature_norm or feature_norm in {"target", "label", "outcome", "classlabel"})
        if row["exact_target_match"] >= 0.995 or (target_named and np.isfinite(row["target_determinism_purity"]) and row["target_determinism_purity"] >= 0.995):
            severe_leakage.append(row)
    if severe_leakage:
        critical.append(
            "Potential direct target leakage detected in features: " + ", ".join(r["feature"] for r in severe_leakage[:10])
        )

    audit = {
        "dataset": key,
        "display_name": DISPLAY_DATASET_NAMES.get(key, key),
        "csv_file": str(path),
        "sha256": hashlib.sha256(path.read_bytes()).hexdigest(),
        "inferred_last_column_target": inferred_target,
        "selected_target": target,
        "target_transform": transform_desc,
        "rows_before_target_cleanup": int(before_drop),
        "rows_used": n_rows,
        "n_features_used": n_features,
        "n_classes": n_unique,
        "min_class_count": min_class,
        "max_class_count": max_class,
        "imbalance_ratio": imbalance_ratio,
        "group_column": group_column or "",
        "n_groups": int(pd.Series(groups).nunique()) if groups is not None else 0,
        "dropped_columns": ", ".join(dropped),
        "n_flagged_leakage_features": len(leakage_rows),
        "critical_issue_count": len(critical),
        "critical_issues": " | ".join(critical),
    }
    for row in leakage_rows:
        row["dataset"] = key
        row["selected_target"] = target

    return PreparedDataset(
        key=key,
        display_name=DISPLAY_DATASET_NAMES.get(key, key),
        path=path,
        df=raw,
        target=target,
        groups=groups,
        group_column=group_column,
        dropped_columns=dropped,
        target_transform=transform_desc,
        audit=audit,
        leakage_rows=leakage_rows,
        critical_issues=critical,
    )


def class_distribution_rows(prepared: PreparedDataset) -> List[Dict[str, Any]]:
    counts = prepared.df[prepared.target].astype(str).value_counts(dropna=False)
    return [
        {
            "dataset": prepared.key,
            "display_name": prepared.display_name,
            "target": prepared.target,
            "class_label": str(label),
            "count": int(count),
            "proportion": float(count / counts.sum()),
        }
        for label, count in counts.items()
    ]


def _split_with_optional_y(splitter: Any, X: pd.DataFrame, y: np.ndarray, groups: Optional[np.ndarray] = None) -> List[Tuple[np.ndarray, np.ndarray]]:
    if groups is not None:
        return list(splitter.split(X, y, groups))
    try:
        return list(splitter.split(X, y))
    except TypeError:
        return list(splitter.split(X))


def make_cv_plan(
    y: np.ndarray,
    seed: int,
    requested_folds: int,
    X: pd.DataFrame,
    groups: Optional[np.ndarray] = None,
    prefer_grouped: bool = False,
    forced_protocol: Optional[str] = None,
) -> CVPlan:
    counts = pd.Series(y).value_counts()
    min_count = int(counts.min()) if len(counts) else 0
    n_classes = int(len(counts))
    requested_folds = int(requested_folds)

    if forced_protocol == "record_stratified":
        n = max(2, min(requested_folds, min_count)) if min_count >= 2 else max(2, min(requested_folds, len(y)))
        if min_count >= 2:
            splitter = StratifiedKFold(n_splits=n, shuffle=True, random_state=seed)
            return CVPlan(f"StratifiedKFold_{n}", n, _split_with_optional_y(splitter, X, y), True, False)
        splitter = KFold(n_splits=n, shuffle=True, random_state=seed)
        return CVPlan(f"KFold_{n}_rare_class_fallback", n, _split_with_optional_y(splitter, X, y), False, False)

    if forced_protocol == "grouped" and groups is not None:
        n_groups = int(pd.Series(groups).nunique())
        n = max(2, min(requested_folds, n_groups, max(min_count, 2)))
        if StratifiedGroupKFold is not None and n_classes > 1:
            splitter = StratifiedGroupKFold(n_splits=n, shuffle=True, random_state=seed)
            return CVPlan(f"StratifiedGroupKFold_{n}", n, _split_with_optional_y(splitter, X, y, groups), True, True)
        splitter = GroupKFold(n_splits=n)
        return CVPlan(f"GroupKFold_{n}", n, _split_with_optional_y(splitter, X, y, groups), False, True)

    if forced_protocol in {"kfold_5", "kfold_10"}:
        n = 5 if forced_protocol.endswith("5") else 10
        n = max(2, min(n, len(y)))
        splitter = KFold(n_splits=n, shuffle=True, random_state=seed)
        return CVPlan(f"KFold_{n}", n, _split_with_optional_y(splitter, X, y), False, False)

    if forced_protocol in {"stratified_5", "stratified_10"}:
        desired = 5 if forced_protocol.endswith("5") else 10
        n = max(2, min(desired, min_count))
        splitter = StratifiedKFold(n_splits=n, shuffle=True, random_state=seed)
        return CVPlan(f"StratifiedKFold_{n}", n, _split_with_optional_y(splitter, X, y), True, False)

    if forced_protocol == "repeated_stratified":
        n = max(2, min(requested_folds, min_count))
        splitter = RepeatedStratifiedKFold(n_splits=n, n_repeats=3, random_state=seed)
        return CVPlan(f"RepeatedStratifiedKFold_{n}x3", n * 3, _split_with_optional_y(splitter, X, y), True, False)

    # Main article protocol.
    if prefer_grouped and groups is not None:
        return make_cv_plan(y, seed, requested_folds, X, groups, forced_protocol="grouped")
    if min_count >= requested_folds and n_classes > 1:
        splitter = StratifiedKFold(n_splits=requested_folds, shuffle=True, random_state=seed)
        return CVPlan(
            f"StratifiedKFold_{requested_folds}", requested_folds,
            _split_with_optional_y(splitter, X, y), True, False,
        )
    if min_count >= 2 and n_classes > 1:
        n = max(2, min(requested_folds, min_count))
        splitter = StratifiedKFold(n_splits=n, shuffle=True, random_state=seed)
        return CVPlan(f"StratifiedKFold_adjusted_{n}", n, _split_with_optional_y(splitter, X, y), True, False)
    n = max(2, min(requested_folds, len(y)))
    splitter = KFold(n_splits=n, shuffle=True, random_state=seed)
    return CVPlan(f"KFold_fallback_{n}_rare_classes", n, _split_with_optional_y(splitter, X, y), False, False)


@dataclass
class ModelSpec:
    name: str
    fit_input: str
    pred_input: str
    factory: Any


def get_model_spec(name: str, seed: int, mlp_max_iter: int = 300) -> ModelSpec:
    if name == "Gaussian NB":
        return ModelSpec(name, "X_cont", "X_cont", lambda: GaussianNB())
    if name == "Bernoulli NB":
        return ModelSpec(name, "X_binary", "X_binary", lambda: BernoulliNB())
    if name == "Categorical NB":
        return ModelSpec(name, "X_disc", "X_disc", lambda: CategoricalNB())
    if name == "WNB":
        return ModelSpec(name, "X_disc", "X_disc", lambda: WeightedCategoricalNB(random_state=seed))
    if name == "NBEM":
        return ModelSpec(name, "data", "data", lambda: NBEMClassifier())
    if name == "Dependency-Aware NBEM":
        return ModelSpec(name, "X_disc", "X_disc", lambda: DependencyAwareNBEM(max_interactions=10, random_state=seed))
    if name == "Adaptive Weighted NBEM":
        return ModelSpec(name, "data", "data", lambda: AdaptiveWeightedNBEM(temperature=1.0, lambda_entropy=0.5))
    if name == "Probabilistic Deep NBEM":
        return ModelSpec(name, "X_cont", "X_cont", lambda: ProbabilisticDeepNBEM(hidden_layer_sizes=(128, 64), max_iter=mlp_max_iter, random_state=seed))
    if name == "Hybrid NBEM":
        return ModelSpec(name, "data", "data", lambda: HybridNBEM(True, True, True, mlp_max_iter=mlp_max_iter, random_state=seed))
    if name == "Hybrid w/o Dependency":
        return ModelSpec(name, "data", "data", lambda: HybridNBEM(False, True, True, mlp_max_iter=mlp_max_iter, random_state=seed))
    if name == "Hybrid w/o Adaptive":
        return ModelSpec(name, "data", "data", lambda: HybridNBEM(True, False, True, mlp_max_iter=mlp_max_iter, random_state=seed))
    if name == "Hybrid w/o Deep":
        return ModelSpec(name, "data", "data", lambda: HybridNBEM(True, True, False, mlp_max_iter=mlp_max_iter, random_state=seed))
    raise ValueError(f"Unknown model name: {name}")


def get_input(data: Dict[str, Any], mode: str) -> Any:
    if mode == "data":
        return data
    if mode == "X_binary":
        return (data["X_disc"] > 0).astype(int)
    return data[mode]


def evaluate_prepared_dataset(
    prepared: PreparedDataset,
    model_names: Sequence[str],
    seed: int,
    requested_folds: int,
    n_bins: int,
    quick: bool,
    prefer_grouped: bool,
    forced_protocol: Optional[str] = None,
    analysis_name: str = "main",
    mlp_max_iter: int = 300,
) -> pd.DataFrame:
    set_reproducibility(seed)
    df = prepared.df.copy()
    groups = prepared.groups.copy() if prepared.groups is not None else None
    if quick and len(df) > 2500:
        sample_idx = df.sample(n=2500, random_state=seed).index.to_numpy()
        df = df.loc[sample_idx].reset_index(drop=True)
        if groups is not None:
            groups = groups[sample_idx]

    encoder = LabelEncoder()
    y = encoder.fit_transform(df[prepared.target].astype(str))
    global_classes = np.arange(len(encoder.classes_))
    plan = make_cv_plan(
        y=y,
        seed=seed,
        requested_folds=requested_folds,
        X=df,
        groups=groups,
        prefer_grouped=prefer_grouped,
        forced_protocol=forced_protocol,
    )

    rows: List[Dict[str, Any]] = []
    for fold, (tr, te) in enumerate(plan.splits, 1):
        train_df = df.iloc[tr].reset_index(drop=True)
        test_df = df.iloc[te].reset_index(drop=True)
        y_train, y_test = y[tr], y[te]
        if len(np.unique(y_train)) < 2:
            for name in model_names:
                rows.append({
                    "analysis": analysis_name, "dataset": prepared.key, "display_name": prepared.display_name,
                    "model": name, "seed": seed, "fold": fold, "cv_protocol": plan.protocol,
                    "n_splits_or_total_folds": plan.n_splits, "status": "SKIPPED: fewer than two training classes",
                })
            continue

        prep = NBEMPreprocessor(n_bins=n_bins).fit(train_df, prepared.target)
        train = prep.transform(train_df)
        test = prep.transform(test_df)
        for name in model_names:
            base = {
                "analysis": analysis_name,
                "dataset": prepared.key,
                "display_name": prepared.display_name,
                "file": str(prepared.path),
                "target": prepared.target,
                "target_transform": prepared.target_transform,
                "model": name,
                "seed": int(seed),
                "fold": int(fold),
                "cv_protocol": plan.protocol,
                "n_splits_or_total_folds": int(plan.n_splits),
                "stratified": bool(plan.stratified),
                "grouped": bool(plan.grouped),
                "n_rows": int(len(df)),
                "n_features": int(df.shape[1] - 1),
                "n_classes": int(len(global_classes)),
                "n_boolean": int(len(prep.bool_cols)),
                "n_categorical": int(len(prep.cat_cols)),
                "n_numerical": int(len(prep.num_cols)),
            }
            try:
                spec = get_model_spec(name, seed=seed, mlp_max_iter=mlp_max_iter)
                model = spec.factory()
                X_train = get_input(train, spec.fit_input)
                X_test = get_input(test, spec.pred_input)
                tracemalloc.start()
                t0 = time.perf_counter()
                model.fit(X_train, y_train)
                train_time = time.perf_counter() - t0
                t1 = time.perf_counter()
                pred = model.predict(X_test)
                raw_proba = model.predict_proba(X_test)
                inference_time = time.perf_counter() - t1
                _, peak = tracemalloc.get_traced_memory()
                tracemalloc.stop()
                model_classes = getattr(model, "classes_", np.unique(y_train))
                proba = align_proba(raw_proba, model_classes, global_classes)
                metrics = compute_metrics(y_test, pred, proba, global_classes)
                rows.append({
                    **base,
                    **metrics,
                    "train_time_s": float(train_time),
                    "inference_time_s": float(inference_time),
                    "peak_memory_mb": float(peak / (1024.0 * 1024.0)),
                    "status": "OK",
                })
            except Exception as exc:
                try:
                    tracemalloc.stop()
                except Exception:
                    pass
                rows.append({
                    **base,
                    **{m: np.nan for m in METRIC_COLUMNS},
                    "status": f"ERROR: {type(exc).__name__}: {exc}",
                })
    return pd.DataFrame(rows)


def experiment_signature(script_path: Path, config_path: Path, args: argparse.Namespace) -> str:
    payload = {
        "script_sha256": hashlib.sha256(script_path.read_bytes()).hexdigest(),
        "config_sha256": hashlib.sha256(config_path.read_bytes()).hexdigest(),
        "folds": args.folds,
        "seeds": args.seeds,
        "quick": args.quick,
        "models": ARTICLE_MODELS,
    }
    return hashlib.sha256(json.dumps(payload, sort_keys=True).encode("utf-8")).hexdigest()[:16]


def safe_filename(value: str) -> str:
    return "".join(ch if ch.isalnum() or ch in "._-" else "_" for ch in value)


def save_dataset_audits(prepared: Sequence[PreparedDataset], output_dir: Path) -> None:
    pd.DataFrame([p.audit for p in prepared]).to_csv(output_dir / "dataset_audit.csv", index=False, encoding="utf-8-sig")
    leakage = [row for p in prepared for row in p.leakage_rows]
    pd.DataFrame(leakage).to_csv(output_dir / "potential_leakage_features.csv", index=False, encoding="utf-8-sig")
    critical_rows = [
        {"dataset": p.key, "issue": issue}
        for p in prepared for issue in p.critical_issues
    ]
    pd.DataFrame(critical_rows).to_csv(output_dir / "critical_audit_issues.csv", index=False, encoding="utf-8-sig")
    classes = [row for p in prepared for row in class_distribution_rows(p)]
    pd.DataFrame(classes).to_csv(output_dir / "dataset_class_distributions.csv", index=False, encoding="utf-8-sig")


def run_main_experiments(
    prepared: Sequence[PreparedDataset],
    output_dir: Path,
    cache_dir: Path,
    seeds: Sequence[int],
    folds: int,
    n_bins: int,
    quick: bool,
    resume: bool,
    mlp_max_iter: int,
    reference_seed: int,
) -> pd.DataFrame:
    frames: List[pd.DataFrame] = []
    total = len(prepared) * len(seeds)
    counter = 0
    for dataset in prepared:
        override_grouped = dataset.groups is not None and dataset.key == "diabetes_130_us_hospitals_for_years_1999_2008"
        for seed in seeds:
            counter += 1
            cache_file = cache_dir / "main" / f"{safe_filename(dataset.key)}__seed_{seed}.csv"
            cache_file.parent.mkdir(parents=True, exist_ok=True)
            if resume and cache_file.exists() and not quick:
                print(f"[{counter}/{total}] Resume: {dataset.key}, seed={seed}")
                frames.append(pd.read_csv(cache_file))
                continue
            models_for_seed = ARTICLE_MODELS if int(seed) == int(reference_seed) else KEY_MODELS
            print(f"[{counter}/{total}] Main: {dataset.key}, seed={seed}, models={len(models_for_seed)}")
            frame = evaluate_prepared_dataset(
                prepared=dataset,
                model_names=models_for_seed,
                seed=int(seed),
                requested_folds=folds,
                n_bins=n_bins,
                quick=quick,
                prefer_grouped=override_grouped,
                analysis_name="main",
                mlp_max_iter=mlp_max_iter,
            )
            frame.to_csv(cache_file, index=False, encoding="utf-8-sig")
            frames.append(frame)
            pd.concat(frames, ignore_index=True).to_csv(output_dir / "all_fold_results_partial.csv", index=False, encoding="utf-8-sig")
    result = pd.concat(frames, ignore_index=True)
    result.to_csv(output_dir / "all_fold_results.csv", index=False, encoding="utf-8-sig")
    return result


def sensitivity_protocols(dataset: PreparedDataset, requested_folds: int) -> List[str]:
    y = dataset.df[dataset.target].astype(str)
    min_count = int(y.value_counts().min())
    protocols: List[str] = []
    if dataset.key == "diabetes_130_us_hospitals_for_years_1999_2008" and dataset.groups is not None:
        protocols.extend(["record_stratified", "grouped"])
    if min_count < requested_folds:
        if min_count >= 2:
            protocols.extend(["stratified_5", "repeated_stratified"])
        else:
            protocols.extend(["kfold_5", "kfold_10"])
    # Explicitly include datasets historically reported with non-standard CV.
    if dataset.key in {"abalone", "cardiotocography", "productivity_prediction_of_garment_employees"}:
        if min_count >= 5:
            protocols.extend(["stratified_5", "stratified_10"])
        else:
            protocols.extend(["kfold_5", "kfold_10"])
    return list(dict.fromkeys(protocols))


def run_sensitivity_experiments(
    prepared: Sequence[PreparedDataset],
    output_dir: Path,
    cache_dir: Path,
    seeds: Sequence[int],
    folds: int,
    n_bins: int,
    quick: bool,
    resume: bool,
    mlp_max_iter: int,
) -> pd.DataFrame:
    frames: List[pd.DataFrame] = []
    jobs: List[Tuple[PreparedDataset, int, str]] = []
    for dataset in prepared:
        for protocol in sensitivity_protocols(dataset, folds):
            for seed in seeds:
                jobs.append((dataset, int(seed), protocol))
    for i, (dataset, seed, protocol) in enumerate(jobs, 1):
        cache_file = cache_dir / "sensitivity" / f"{safe_filename(dataset.key)}__{protocol}__seed_{seed}.csv"
        cache_file.parent.mkdir(parents=True, exist_ok=True)
        if resume and cache_file.exists() and not quick:
            print(f"[Sensitivity {i}/{len(jobs)}] Resume: {dataset.key}, {protocol}, seed={seed}")
            frames.append(pd.read_csv(cache_file))
            continue
        print(f"[Sensitivity {i}/{len(jobs)}] {dataset.key}, {protocol}, seed={seed}")
        frame = evaluate_prepared_dataset(
            prepared=dataset,
            model_names=KEY_MODELS,
            seed=seed,
            requested_folds=folds,
            n_bins=n_bins,
            quick=quick,
            prefer_grouped=False,
            forced_protocol=protocol,
            analysis_name=f"sensitivity:{protocol}",
            mlp_max_iter=mlp_max_iter,
        )
        frame.to_csv(cache_file, index=False, encoding="utf-8-sig")
        frames.append(frame)
    if frames:
        result = pd.concat(frames, ignore_index=True)
    else:
        result = pd.DataFrame()
    result.to_csv(output_dir / "cv_sensitivity_all_fold_results.csv", index=False, encoding="utf-8-sig")
    return result


def holm_adjust(p_values: Sequence[float]) -> np.ndarray:
    p = np.asarray(p_values, dtype=float)
    adjusted = np.full_like(p, np.nan)
    valid_idx = np.where(np.isfinite(p))[0]
    if len(valid_idx) == 0:
        return adjusted
    order = valid_idx[np.argsort(p[valid_idx])]
    running = 0.0
    m = len(order)
    for rank, idx in enumerate(order):
        value = min(1.0, (m - rank) * p[idx])
        running = max(running, value)
        adjusted[idx] = running
    return adjusted


def summarize_fold_results(results: pd.DataFrame, output_dir: Path, reference_seed: int) -> Dict[str, pd.DataFrame]:
    ok = results[results["status"] == "OK"].copy()
    if ok.empty:
        raise RuntimeError("No successful model-fold results were produced.")

    # Fold -> dataset/model/seed.
    seed_summary = ok.groupby(["dataset", "display_name", "model", "seed"], as_index=False)[METRIC_COLUMNS].mean()
    seed_summary.to_csv(output_dir / "summary_dataset_model_seed.csv", index=False, encoding="utf-8-sig")

    # Across seeds, per dataset/model.
    agg_map = {m: ["mean", "std"] for m in METRIC_COLUMNS}
    dataset_model = seed_summary.groupby(["dataset", "display_name", "model"])[METRIC_COLUMNS].agg(agg_map).reset_index()
    dataset_model.columns = ["_".join([str(x) for x in col if str(x)]) for col in dataset_model.columns]
    dataset_model = dataset_model.rename(columns={"dataset_": "dataset", "display_name_": "display_name", "model_": "model"})
    dataset_model.to_csv(output_dir / "summary_dataset_model.csv", index=False, encoding="utf-8-sig")

    # Dataset-balanced overall metric for each seed, then mean +/- SD across seeds.
    per_seed_model = seed_summary.groupby(["model", "seed"], as_index=False)[METRIC_COLUMNS].mean()
    overall = per_seed_model.groupby("model")[METRIC_COLUMNS].agg(["mean", "std"]).reset_index()
    overall.columns = ["_".join([str(x) for x in col if str(x)]) for col in overall.columns]
    overall = overall.rename(columns={"model_": "model"})
    overall = overall.sort_values("f1_weighted_mean", ascending=False)
    overall.to_csv(output_dir / "average_metrics_by_model.csv", index=False, encoding="utf-8-sig")
    per_seed_model.to_csv(output_dir / "average_metrics_by_model_seed.csv", index=False, encoding="utf-8-sig")

    # Datasetwise pivots, averaged over seeds.
    dataset_mean = seed_summary.groupby(["dataset", "display_name", "model"], as_index=False)[METRIC_COLUMNS].mean()
    pivots: Dict[str, pd.DataFrame] = {}
    for metric in ["f1_weighted", "f1_macro", "precision_weighted", "recall_weighted", "accuracy", "roc_auc_weighted"]:
        pivot = dataset_mean.pivot(index=["dataset", "display_name"], columns="model", values=metric)
        pivot.to_csv(output_dir / f"datasetwise_{metric}_pivot.csv", encoding="utf-8-sig")
        pivots[metric] = pivot

    # All-model ranks use the common reference seed so every model is evaluated on the same splits.
    reference_dataset = seed_summary[seed_summary["seed"] == int(reference_seed)].copy()
    reference_overall = reference_dataset.groupby("model", as_index=False)[METRIC_COLUMNS].mean()
    reference_overall = reference_overall.rename(columns={m: f"{m}_mean" for m in METRIC_COLUMNS})
    reference_overall.to_csv(output_dir / "reference_seed_average_metrics_by_model.csv", index=False, encoding="utf-8-sig")
    reference_pivots: Dict[str, pd.DataFrame] = {}
    rank_outputs: List[pd.DataFrame] = []
    for metric in ["f1_weighted", "f1_macro"]:
        pivot = reference_dataset.pivot(index=["dataset", "display_name"], columns="model", values=metric)
        reference_pivots[metric] = pivot
        pivot.to_csv(output_dir / f"reference_seed_{metric}_pivot.csv", encoding="utf-8-sig")
        ranks = pivot.rank(axis=1, ascending=False, method="average")
        avg_rank = ranks.mean(axis=0).sort_values().reset_index()
        avg_rank.columns = ["model", f"average_rank_{metric}"]
        avg_rank.to_csv(output_dir / f"average_rank_{metric}.csv", index=False, encoding="utf-8-sig")
        if not rank_outputs:
            rank_outputs.append(avg_rank)
        else:
            rank_outputs[0] = rank_outputs[0].merge(avg_rank, on="model", how="outer")
    ranks_combined = rank_outputs[0]
    ranks_combined.to_csv(output_dir / "average_ranks_combined.csv", index=False, encoding="utf-8-sig")

    # Win/tie/loss for weighted and macro F1.
    wtl_rows: List[Dict[str, Any]] = []
    for metric in ["f1_weighted", "f1_macro"]:
        pivot = pivots[metric]
        for baseline in ["NBEM", "Hybrid NBEM"]:
            if baseline not in pivot.columns:
                continue
            for model in pivot.columns:
                if model == baseline:
                    continue
                diff = pivot[model] - pivot[baseline]
                wtl_rows.append({
                    "metric": metric,
                    "comparison": f"{model} vs {baseline}",
                    "model": model,
                    "baseline": baseline,
                    "wins": int((diff > 1e-12).sum()),
                    "ties": int((diff.abs() <= 1e-12).sum()),
                    "losses": int((diff < -1e-12).sum()),
                    "mean_difference": float(diff.mean()),
                })
    wtl = pd.DataFrame(wtl_rows)
    wtl.to_csv(output_dir / "win_tie_loss.csv", index=False, encoding="utf-8-sig")

    statistical = make_statistical_tests(pivots, reference_pivots)
    statistical["wilcoxon"].to_csv(output_dir / "wilcoxon_pairwise_holm.csv", index=False, encoding="utf-8-sig")
    statistical["friedman"].to_csv(output_dir / "friedman_tests.csv", index=False, encoding="utf-8-sig")

    return {
        "ok": ok,
        "seed_summary": seed_summary,
        "dataset_model": dataset_model,
        "overall": overall,
        "per_seed_model": per_seed_model,
        "dataset_mean": dataset_mean,
        "pivots": pivots,
        "reference_pivots": reference_pivots,
        "reference_overall": reference_overall,
        "ranks": ranks_combined,
        "win_tie_loss": wtl,
        "wilcoxon": statistical["wilcoxon"],
        "friedman": statistical["friedman"],
    }


def make_statistical_tests(pivots: Dict[str, pd.DataFrame], reference_pivots: Dict[str, pd.DataFrame]) -> Dict[str, pd.DataFrame]:
    pair_rows: List[Dict[str, Any]] = []
    friedman_rows: List[Dict[str, Any]] = []
    comparators = ["WNB", "NBEM", "Adaptive Weighted NBEM"]
    for metric in ["f1_weighted", "f1_macro"]:
        pivot = pivots[metric]
        metric_start = len(pair_rows)
        for comparator in comparators:
            if comparator not in pivot.columns or "Hybrid NBEM" not in pivot.columns:
                continue
            tmp = pivot[[comparator, "Hybrid NBEM"]].dropna()
            stat = pval = np.nan
            if wilcoxon is not None and len(tmp) >= 2:
                try:
                    stat, pval = wilcoxon(tmp[comparator], tmp["Hybrid NBEM"], zero_method="wilcox", alternative="two-sided")
                except Exception:
                    pass
            pair_rows.append({
                "metric": metric,
                "comparator": comparator,
                "baseline": "Hybrid NBEM",
                "n_datasets": int(len(tmp)),
                "mean_comparator": float(tmp[comparator].mean()) if len(tmp) else np.nan,
                "mean_hybrid": float(tmp["Hybrid NBEM"].mean()) if len(tmp) else np.nan,
                "mean_difference_comparator_minus_hybrid": float((tmp[comparator] - tmp["Hybrid NBEM"]).mean()) if len(tmp) else np.nan,
                "wilcoxon_statistic": float(stat) if np.isfinite(stat) else np.nan,
                "p_raw": float(pval) if np.isfinite(pval) else np.nan,
            })
        metric_end = len(pair_rows)
        if metric_end > metric_start:
            adjusted = holm_adjust([row["p_raw"] for row in pair_rows[metric_start:metric_end]])
            for row, padj in zip(pair_rows[metric_start:metric_end], adjusted):
                row["p_holm"] = float(padj) if np.isfinite(padj) else np.nan
                row["significant_holm_0_05"] = bool(padj < 0.05) if np.isfinite(padj) else False

        valid = reference_pivots[metric].dropna(axis=0, how="any")
        if friedmanchisquare is not None and valid.shape[0] >= 2 and valid.shape[1] >= 3:
            try:
                stat, pval = friedmanchisquare(*[valid[c].values for c in valid.columns])
                friedman_rows.append({
                    "metric": metric,
                    "n_datasets_complete": int(valid.shape[0]),
                    "n_models": int(valid.shape[1]),
                    "friedman_statistic": float(stat),
                    "p_value": float(pval),
                    "significant_0_05": bool(pval < 0.05),
                    "models_included": ", ".join(valid.columns),
                })
            except Exception as exc:
                friedman_rows.append({"metric": metric, "error": str(exc)})
    return {"wilcoxon": pd.DataFrame(pair_rows), "friedman": pd.DataFrame(friedman_rows)}


def summarize_sensitivity(results: pd.DataFrame, output_dir: Path) -> pd.DataFrame:
    if results.empty:
        summary = pd.DataFrame()
        summary.to_csv(output_dir / "cv_sensitivity_summary.csv", index=False, encoding="utf-8-sig")
        return summary
    ok = results[results["status"] == "OK"].copy()
    seed_level = ok.groupby(["dataset", "display_name", "analysis", "cv_protocol", "model", "seed"], as_index=False)[
        ["f1_weighted", "f1_macro", "accuracy", "balanced_accuracy"]
    ].mean()
    summary = seed_level.groupby(["dataset", "display_name", "analysis", "cv_protocol", "model"])[
        ["f1_weighted", "f1_macro", "accuracy", "balanced_accuracy"]
    ].agg(["mean", "std"]).reset_index()
    summary.columns = ["_".join([str(x) for x in col if str(x)]) for col in summary.columns]
    summary = summary.rename(columns={
        "dataset_": "dataset", "display_name_": "display_name", "analysis_": "analysis",
        "cv_protocol_": "cv_protocol", "model_": "model",
    })
    summary.to_csv(output_dir / "cv_sensitivity_summary.csv", index=False, encoding="utf-8-sig")
    return summary


def create_article_tables(analysis: Dict[str, pd.DataFrame], sensitivity: pd.DataFrame, audits: pd.DataFrame, output_dir: Path) -> Dict[str, pd.DataFrame]:
    tables_dir = output_dir / "article_tables"
    tables_dir.mkdir(parents=True, exist_ok=True)
    overall = analysis["overall"]
    dataset_mean = analysis["dataset_mean"]
    key = overall[overall["model"].isin(KEY_MODELS)].copy()
    key_order = {m: i for i, m in enumerate(KEY_MODELS)}
    key["order"] = key["model"].map(key_order)
    key = key.sort_values("order").drop(columns=["order"])
    key.to_csv(tables_dir / "table_main_average_performance.csv", index=False, encoding="utf-8-sig")

    outputs: Dict[str, pd.DataFrame] = {"main": key}
    for metric in ["f1_weighted", "f1_macro"]:
        p = dataset_mean[dataset_mean["model"].isin(KEY_MODELS)].pivot(
            index=["dataset", "display_name"], columns="model", values=metric
        ).reset_index()
        ordered = ["dataset", "display_name"] + [m for m in KEY_MODELS if m in p.columns]
        p = p[ordered]
        p.to_csv(tables_dir / f"table_datasetwise_{metric}.csv", index=False, encoding="utf-8-sig")
        outputs[metric] = p

    pr = dataset_mean[dataset_mean["model"].isin(KEY_MODELS)].copy()
    pr["precision_recall_weighted"] = pr.apply(
        lambda r: f"{r['precision_weighted']:.3f}({r['recall_weighted']:.3f})", axis=1
    )
    pr_pivot = pr.pivot(index=["dataset", "display_name"], columns="model", values="precision_recall_weighted").reset_index()
    pr_pivot.to_csv(tables_dir / "table_datasetwise_weighted_precision_recall.csv", index=False, encoding="utf-8-sig")
    outputs["precision_recall"] = pr_pivot

    reference_overall = analysis["reference_overall"]
    ablation = reference_overall[reference_overall["model"].isin(ABLATION_MODELS)].copy()
    hybrid_value = float(reference_overall.loc[reference_overall["model"] == "Hybrid NBEM", "f1_weighted_mean"].iloc[0])
    ablation["delta_f1_weighted_vs_hybrid"] = ablation["f1_weighted_mean"] - hybrid_value
    ablation = ablation.sort_values("f1_weighted_mean", ascending=False)
    ablation.to_csv(tables_dir / "table_ablation.csv", index=False, encoding="utf-8-sig")
    outputs["ablation"] = ablation

    analysis["ranks"].to_csv(tables_dir / "table_average_ranks.csv", index=False, encoding="utf-8-sig")
    analysis["wilcoxon"].to_csv(tables_dir / "table_wilcoxon_holm.csv", index=False, encoding="utf-8-sig")
    analysis["friedman"].to_csv(tables_dir / "table_friedman.csv", index=False, encoding="utf-8-sig")

    cost_cols = ["model", "train_time_s_mean", "train_time_s_std", "inference_time_s_mean", "inference_time_s_std", "peak_memory_mb_mean", "peak_memory_mb_std"]
    cost = key[[c for c in cost_cols if c in key.columns]].copy()
    cost.to_csv(tables_dir / "table_computational_cost.csv", index=False, encoding="utf-8-sig")
    outputs["cost"] = cost

    diabetes = dataset_mean[
        (dataset_mean["dataset"] == "diabetes_130_us_hospitals_for_years_1999_2008")
        & (dataset_mean["model"].isin(KEY_MODELS))
    ].copy()
    diabetes.to_csv(tables_dir / "table_diabetes_corrected_results.csv", index=False, encoding="utf-8-sig")
    outputs["diabetes"] = diabetes

    sensitivity.to_csv(tables_dir / "table_cv_sensitivity.csv", index=False, encoding="utf-8-sig")
    audits.to_csv(tables_dir / "table_dataset_audit_summary.csv", index=False, encoding="utf-8-sig")

    limitations = pd.DataFrame([
        {"Threat or limitation": "Benchmark scope", "Mitigation / interpretation": "Claims are restricted to the NB/NBEM family; modern ensemble and tabular-foundation baselines are left to future work."},
        {"Threat or limitation": "Non-cross-fitted stacker", "Mitigation / interpretation": "Theorem 3.3 is an idealized theoretical result and is not presented as a finite-sample guarantee for the current implementation."},
        {"Threat or limitation": "Seed sensitivity", "Mitigation / interpretation": "Five predefined random seeds are reported using mean and standard deviation."},
        {"Threat or limitation": "Class imbalance", "Mitigation / interpretation": "Macro-F1 is reported together with weighted-F1, accuracy, and balanced accuracy."},
        {"Threat or limitation": "CV feasibility for rare classes", "Mitigation / interpretation": "Adjusted or non-stratified protocols are explicitly reported and subjected to a separate sensitivity analysis."},
        {"Threat or limitation": "Heuristic feature typing", "Mitigation / interpretation": "Features with at most two observed values are treated as Boolean; this may misclassify discrete numerical variables in small samples."},
    ])
    limitations.to_csv(tables_dir / "table_threats_to_validity.csv", index=False, encoding="utf-8-sig")
    outputs["limitations"] = limitations

    novelty = pd.DataFrame([
        {"Aspect": "Component weighting", "Original NBEM": "Fixed feature-proportion weights", "This work": "Instance-dependent entropy/margin adaptive weights"},
        {"Aspect": "Feature dependence", "Original NBEM": "Conditional-independence components", "This work": "Dependency-aware interaction component"},
        {"Aspect": "Nonlinear representation", "Original NBEM": "Not included", "This work": "Probabilistic MLP component"},
        {"Aspect": "Fusion", "Original NBEM": "Direct weighted posterior combination", "This work": "Hybrid probability stacking"},
        {"Aspect": "Empirical analysis", "Original NBEM": "Base-family comparison", "This work": "Multi-seed, macro-F1, ablation, and CV-sensitivity analysis"},
    ])
    novelty.to_csv(tables_dir / "table_nbem_vs_proposed_method.csv", index=False, encoding="utf-8-sig")
    outputs["novelty"] = novelty
    return outputs


def latex_escape(value: Any) -> str:
    text = str(value)
    replacements = {
        "\\": r"\textbackslash{}", "&": r"\&", "%": r"\%", "$": r"\$", "#": r"\#",
        "_": r"\_", "{": r"\{", "}": r"\}", "~": r"\textasciitilde{}", "^": r"\textasciicircum{}",
    }
    return "".join(replacements.get(ch, ch) for ch in text)


def write_latex_table(df: pd.DataFrame, path: Path, caption: str, label: str, float_fmt: str = ".4f") -> None:
    if df.empty:
        path.write_text("% No data were available for this table.\n", encoding="utf-8")
        return
    cols = list(df.columns)
    align = "l" + "r" * (len(cols) - 1)
    lines = [
        r"\begin{table}[!htbp]", r"\centering", r"\scriptsize",
        rf"\caption{{{caption}}}", rf"\label{{{label}}}",
        rf"\begin{{adjustbox}}{{max width=\textwidth}}", rf"\begin{{tabular}}{{{align}}}", r"\toprule",
        " & ".join(latex_escape(c) for c in cols) + r" \\", r"\midrule",
    ]
    for _, row in df.iterrows():
        values = []
        for value in row:
            if pd.isna(value):
                values.append("--")
            elif isinstance(value, (float, np.floating)):
                values.append(format(float(value), float_fmt))
            else:
                values.append(latex_escape(value))
        lines.append(" & ".join(values) + r" \\")
    lines += [r"\bottomrule", r"\end{tabular}", r"\end{adjustbox}", r"\end{table}"]
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def create_latex_tables(tables: Dict[str, pd.DataFrame], output_dir: Path) -> None:
    latex_dir = output_dir / "article_tables" / "latex"
    latex_dir.mkdir(parents=True, exist_ok=True)
    write_latex_table(tables["main"], latex_dir / "table_main_average_performance.tex",
                      "Average predictive performance across the benchmark datasets, reported over five predefined seeds.", "tab:main_average")
    write_latex_table(tables["f1_weighted"], latex_dir / "table_datasetwise_weighted_f1.tex",
                      "Dataset-wise weighted F1-score comparison within the NB/NBEM family.", "tab:dataset_weighted_f1")
    write_latex_table(tables["f1_macro"], latex_dir / "table_datasetwise_macro_f1.tex",
                      "Dataset-wise macro-F1 comparison within the NB/NBEM family.", "tab:dataset_macro_f1")
    write_latex_table(tables["precision_recall"], latex_dir / "table_datasetwise_precision_recall.tex",
                      "Dataset-wise weighted Precision and Recall, reported as Precision(Recall).", "tab:dataset_pr")
    write_latex_table(tables["ablation"], latex_dir / "table_ablation.tex",
                      "Complete ablation study of Hybrid NBEM.", "tab:ablation")
    write_latex_table(tables["cost"], latex_dir / "table_computational_cost.tex",
                      "Empirical computational cost averaged across completed folds and seeds.", "tab:cost")
    write_latex_table(tables["diabetes"], latex_dir / "table_diabetes_results.tex",
                      "Corrected Diabetes results under the explicitly reported validation protocol.", "tab:diabetes")
    write_latex_table(tables["limitations"], latex_dir / "table_threats_to_validity.tex",
                      "Structured threats to validity and their interpretation.", "tab:threats")
    write_latex_table(tables["novelty"], latex_dir / "table_nbem_vs_proposed.tex",
                      "Direct comparison between the original NBEM and the proposed extension.", "tab:novelty")


def make_article_figures(analysis: Dict[str, pd.DataFrame], output_dir: Path) -> None:
    fig_dir = output_dir / "article_figures"
    fig_dir.mkdir(parents=True, exist_ok=True)
    dataset_mean = analysis["dataset_mean"]
    overall = analysis["overall"]

    # Figure 1: dataset-wise weighted F1 for the four principal within-family methods.
    key = dataset_mean[dataset_mean["model"].isin(KEY_MODELS)].pivot(
        index="display_name", columns="model", values="f1_weighted"
    )
    key = key[[m for m in KEY_MODELS if m in key.columns]]
    plt.figure(figsize=(15, 6.5))
    for model in key.columns:
        plt.plot(range(len(key)), key[model].to_numpy(), marker="o", linewidth=1.4, markersize=3.5, label=model)
    plt.xticks(range(len(key)), key.index, rotation=75, ha="right")
    plt.ylabel("Weighted F1-score")
    plt.xlabel("Dataset")
    plt.title("Dataset-wise weighted F1-score comparison within the NB/NBEM family")
    plt.legend(fontsize=8, ncol=2)
    plt.tight_layout()
    plt.savefig(fig_dir / "figure1_datasetwise_weighted_f1.png", dpi=300, bbox_inches="tight")
    plt.close()

    # Figure 2: average weighted/macro metrics with seed standard deviations.
    key_overall = overall[overall["model"].isin(KEY_MODELS)].copy()
    key_overall["order"] = key_overall["model"].map({m: i for i, m in enumerate(KEY_MODELS)})
    key_overall = key_overall.sort_values("order")
    metrics = [
        ("f1_weighted", "Weighted F1"),
        ("f1_macro", "Macro-F1"),
        ("precision_weighted", "Weighted Precision"),
        ("recall_weighted", "Weighted Recall"),
    ]
    x = np.arange(len(key_overall))
    width = 0.19
    plt.figure(figsize=(11, 5.8))
    for j, (metric, label) in enumerate(metrics):
        means = key_overall[f"{metric}_mean"].to_numpy()
        stds = key_overall[f"{metric}_std"].fillna(0).to_numpy()
        plt.bar(x + (j - 1.5) * width, means, width, yerr=stds, capsize=2, label=label)
    plt.xticks(x, key_overall["model"], rotation=18, ha="right")
    plt.ylabel("Average score")
    plt.ylim(0, 1.0)
    plt.title("Average performance across datasets and predefined seeds")
    plt.legend(fontsize=8)
    plt.tight_layout()
    plt.savefig(fig_dir / "figure2_average_performance.png", dpi=300, bbox_inches="tight")
    plt.close()

    # Figure 3: ablation with both weighted and macro F1.
    abl = analysis["reference_overall"]
    abl = abl[abl["model"].isin(ABLATION_MODELS)].copy().sort_values("f1_weighted_mean", ascending=True)
    y = np.arange(len(abl))
    plt.figure(figsize=(10.5, 6.2))
    plt.barh(y - 0.18, abl["f1_weighted_mean"], height=0.35, label="Weighted F1")
    plt.barh(y + 0.18, abl["f1_macro_mean"], height=0.35, label="Macro-F1")
    plt.yticks(y, abl["model"])
    plt.xlabel("Average score")
    plt.title("Ablation study of Hybrid NBEM components")
    plt.legend(fontsize=8)
    plt.tight_layout()
    plt.savefig(fig_dir / "figure3_ablation_weighted_macro_f1.png", dpi=300, bbox_inches="tight")
    plt.close()


def make_excel_report(output_dir: Path) -> None:
    xlsx = output_dir / "NBEM_article_results.xlsx"
    files = sorted([p for p in output_dir.glob("*.csv")]) + sorted((output_dir / "article_tables").glob("*.csv"))
    try:
        with pd.ExcelWriter(xlsx, engine="openpyxl") as writer:
            used: set[str] = set()
            for path in files:
                try:
                    df = pd.read_csv(path)
                except Exception:
                    continue
                base = re.sub(r"[^A-Za-z0-9_]", "_", path.stem)[:31] or "sheet"
                sheet = base
                n = 1
                while sheet in used:
                    suffix = f"_{n}"
                    sheet = (base[:31-len(suffix)] + suffix)
                    n += 1
                used.add(sheet)
                df.to_excel(writer, sheet_name=sheet, index=False)
        # Basic professional formatting.
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        wb = load_workbook(xlsx)
        for ws in wb.worksheets:
            ws.freeze_panes = "A2"
            ws.sheet_view.showGridLines = False
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1F4E78")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            for col in ws.columns:
                letter = col[0].column_letter
                max_len = min(45, max(len(str(c.value)) if c.value is not None else 0 for c in col) + 2)
                ws.column_dimensions[letter].width = max(10, max_len)
        wb.save(xlsx)
    except Exception as exc:
        print(f"Excel report could not be created: {exc}")


def save_environment_report(output_dir: Path, args: argparse.Namespace, signature: str, config: Dict[str, Any]) -> None:
    packages: Dict[str, str] = {}
    for mod_name in ["numpy", "pandas", "sklearn", "scipy", "matplotlib", "openpyxl"]:
        try:
            mod = __import__(mod_name)
            packages[mod_name] = getattr(mod, "__version__", "unknown")
        except Exception:
            packages[mod_name] = "not installed"
    report = {
        "experiment_signature": signature,
        "python": sys.version,
        "platform": platform.platform(),
        "processor": platform.processor(),
        "machine": platform.machine(),
        "arguments": vars(args),
        "packages": packages,
        "methodological_notes": {
            "scope": "NB and NBEM family only; no modern ensemble or tabular-foundation baseline is included.",
            "cross_validation": "Stratified 10-fold CV is used when feasible; adjusted, grouped, or KFold protocols are explicitly recorded per row.",
            "data_leakage_prevention": "Imputation, encoding, scaling, discretization, feature weighting, and interaction selection are fitted only on the training fold.",
            "diabetes": "The configured target is readmitted. Patient-grouped CV is used when patient_nbr is available; record-level CV is retained only as a sensitivity comparator.",
            "productivity": "The continuous actual_productivity target is converted using the fixed, predeclared bins in dataset_config.json.",
            "feature_typing_limitation": "At most two observed values implies Boolean; this heuristic may misclassify discrete numerical variables in small samples.",
            "cross_fitting_limitation": "The current Hybrid NBEM stacker is not fully cross-fitted; Theorem 3.3 describes an idealized theoretical framework, not a finite-sample guarantee for this implementation.",
        },
        "config": config,
    }
    with open(output_dir / "execution_environment.json", "w", encoding="utf-8") as f:
        json.dump(report, f, indent=2, ensure_ascii=False)


def validate_final_outputs(
    results: pd.DataFrame,
    sensitivity_results: pd.DataFrame,
    prepared: Sequence[PreparedDataset],
    seeds: Sequence[int],
    reference_seed: int,
    output_dir: Path,
) -> Dict[str, Any]:
    ok = results[results["status"] == "OK"].copy()
    errors = results[results["status"] != "OK"].copy()
    expected = set()
    for dataset in prepared:
        for seed in seeds:
            models = ARTICLE_MODELS if int(seed) == int(reference_seed) else KEY_MODELS
            expected.update((dataset.key, model, int(seed)) for model in models)
    observed = set(zip(ok["dataset"], ok["model"], ok["seed"].astype(int)))
    missing = sorted(expected - observed)
    key_errors = errors[errors["model"].isin(KEY_MODELS)] if not errors.empty else errors
    sensitivity_errors = (
        sensitivity_results[sensitivity_results["status"] != "OK"]
        if not sensitivity_results.empty and "status" in sensitivity_results.columns else pd.DataFrame()
    )
    figure_files = sorted((output_dir / "article_figures").glob("*.png"))
    report = {
        "passed": bool(
            len(missing) == 0
            and len(key_errors) == 0
            and (ok["f1_macro"].notna().all() if not ok.empty else False)
            and "WNB" in set(ok["model"])
            and len(figure_files) == 3
        ),
        "n_expected_dataset_model_seed_combinations": len(expected),
        "n_observed_dataset_model_seed_combinations": len(observed),
        "n_missing_combinations": len(missing),
        "missing_combinations_first_50": [list(x) for x in missing[:50]],
        "n_all_non_ok_rows": int(len(errors)),
        "n_key_model_non_ok_rows": int(len(key_errors)),
        "n_sensitivity_non_ok_rows": int(len(sensitivity_errors)),
        "macro_f1_complete": bool(ok["f1_macro"].notna().all()) if not ok.empty else False,
        "wnb_present": bool("WNB" in set(ok["model"])),
        "article_figure_files": [p.name for p in figure_files],
        "exactly_three_article_figures": bool(len(figure_files) == 3),
    }
    with open(output_dir / "FINAL_VALIDATION.json", "w", encoding="utf-8") as f:
        json.dump(report, f, indent=2, ensure_ascii=False)
    errors.to_csv(output_dir / "non_ok_model_rows.csv", index=False, encoding="utf-8-sig")
    sensitivity_errors.to_csv(output_dir / "non_ok_sensitivity_rows.csv", index=False, encoding="utf-8-sig")
    return report


def write_run_summary(output_dir: Path, analysis: Dict[str, pd.DataFrame], audits: pd.DataFrame, sensitivity: pd.DataFrame) -> None:
    overall = analysis["overall"]
    lines = [
        "NBEM final article experiment run summary",
        "========================================",
        "",
        "Scope: NB/NBEM-family comparison only. No boosting or tabular-foundation baselines were executed.",
        "",
        f"Datasets audited: {len(audits)}",
        f"Critical audit issues after configuration: {int((audits['critical_issue_count'] > 0).sum()) if not audits.empty else 0}",
        f"Sensitivity summary rows: {len(sensitivity)}",
        "",
        "Key model averages (dataset-balanced, mean across seeds):",
    ]
    for model in KEY_MODELS:
        row = overall[overall["model"] == model]
        if row.empty:
            continue
        r = row.iloc[0]
        lines.append(
            f"- {model}: weighted-F1={r['f1_weighted_mean']:.4f} +/- {r['f1_weighted_std']:.4f}; "
            f"macro-F1={r['f1_macro_mean']:.4f} +/- {r['f1_macro_std']:.4f}"
        )
    lines += [
        "",
        "Important interpretation:",
        "- Claims must remain restricted to the NB/NBEM family.",
        "- The Hybrid NBEM stacker in this implementation is not fully cross-fitted.",
        "- Figures 4 and 5 are intentionally not generated; computational cost is provided as a table only.",
    ]
    (output_dir / "RUN_SUMMARY.txt").write_text("\n".join(lines) + "\n", encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Run the final NB/NBEM-family article experiments.")
    parser.add_argument("--project-root", type=str, default=".")
    parser.add_argument("--config", type=str, default="config/dataset_config.json")
    parser.add_argument("--output-dir", type=str, default="results/final_article_run")
    parser.add_argument("--folds", type=int, default=None, help="Override requested folds in config.")
    parser.add_argument("--seeds", type=int, nargs="*", default=None, help="Override seeds in config.")
    parser.add_argument("--datasets", nargs="*", default=None, help="Optional dataset-folder filter.")
    parser.add_argument("--quick", action="store_true", help="Pipeline validation only; never use quick results in the article.")
    parser.add_argument("--no-resume", action="store_true")
    parser.add_argument("--audit-only", action="store_true")
    parser.add_argument("--allow-critical-audit-issues", action="store_true", help="Not recommended; continue despite critical target/leakage audit findings.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    project_root = Path(args.project_root).resolve()
    config_path = Path(args.config)
    if not config_path.is_absolute():
        config_path = project_root / config_path
    config = load_json(config_path)
    output_dir = Path(args.output_dir)
    if not output_dir.is_absolute():
        output_dir = project_root / output_dir
    output_dir.mkdir(parents=True, exist_ok=True)

    seeds = args.seeds or config.get("seeds", [13, 21, 42, 87, 123])
    folds = int(args.folds or config.get("requested_folds", 10))
    n_bins = int(config.get("n_bins", 5))
    mlp_max_iter = int(config.get("mlp_max_iter", 300 if not args.quick else 60))
    reference_seed = int(config.get("reference_seed", 42))
    if reference_seed not in [int(x) for x in seeds]:
        raise ValueError("reference_seed must be included in the seed list.")
    args.seeds = [int(x) for x in seeds]
    args.folds = folds

    paths = find_datasets(project_root, use_canonical_20=bool(config.get("use_canonical_20", True)))
    if args.datasets:
        filters = {normalize_name(x) for x in args.datasets}
        paths = [p for p in paths if normalize_name(canonical_dataset_key(p)) in filters or normalize_name(p.parent.name) in filters]
    if not paths:
        raise FileNotFoundError(f"No dataset CSV files were found under {project_root / 'datasets' / 'processed'}")

    prepared: List[PreparedDataset] = []
    print("Auditing and preparing datasets...")
    for i, path in enumerate(paths, 1):
        print(f"[Audit {i}/{len(paths)}] {path.parent.name}")
        prepared.append(prepare_dataset(path, config))
    save_dataset_audits(prepared, output_dir)
    audits_df = pd.DataFrame([p.audit for p in prepared])

    critical = [(p.key, issue) for p in prepared for issue in p.critical_issues]
    if critical and not args.allow_critical_audit_issues:
        message = [
            "Critical dataset-audit issues were found. The long experiment was stopped to prevent invalid article results.",
            "Review results/final_article_run/critical_audit_issues.csv and config/dataset_config.json.",
            "Issues:",
        ] + [f"- {dataset}: {issue}" for dataset, issue in critical]
        (output_dir / "CRITICAL_ACTION_REQUIRED.txt").write_text("\n".join(message) + "\n", encoding="utf-8")
        raise RuntimeError("\n".join(message))
    if args.audit_only:
        print(f"Audit completed. Results: {output_dir}")
        return

    script_path = Path(__file__).resolve()
    signature = experiment_signature(script_path, config_path, args)
    cache_dir = output_dir / "cache" / signature
    save_environment_report(output_dir, args, signature, config)

    results = run_main_experiments(
        prepared=prepared,
        output_dir=output_dir,
        cache_dir=cache_dir,
        seeds=seeds,
        folds=folds,
        n_bins=n_bins,
        quick=args.quick,
        resume=not args.no_resume,
        mlp_max_iter=mlp_max_iter,
        reference_seed=reference_seed,
    )
    analysis = summarize_fold_results(results, output_dir, reference_seed=reference_seed)

    sensitivity_results = run_sensitivity_experiments(
        prepared=prepared,
        output_dir=output_dir,
        cache_dir=cache_dir,
        seeds=seeds,
        folds=folds,
        n_bins=n_bins,
        quick=args.quick,
        resume=not args.no_resume,
        mlp_max_iter=mlp_max_iter,
    )
    sensitivity_summary = summarize_sensitivity(sensitivity_results, output_dir)
    tables = create_article_tables(analysis, sensitivity_summary, audits_df, output_dir)
    create_latex_tables(tables, output_dir)
    make_article_figures(analysis, output_dir)
    make_excel_report(output_dir)
    validation = validate_final_outputs(results, sensitivity_results, prepared, seeds, reference_seed, output_dir)
    write_run_summary(output_dir, analysis, audits_df, sensitivity_summary)

    # Copy the exact configuration and script used into the result folder.
    shutil.copy2(config_path, output_dir / "dataset_config_used.json")
    shutil.copy2(script_path, output_dir / "nbem_article_experiments_v3_used.py")
    if not validation["passed"]:
        raise RuntimeError(
            "The experiment completed, but final validation failed. Review FINAL_VALIDATION.json, "
            "non_ok_model_rows.csv, and non_ok_sensitivity_rows.csv before using article results."
        )
    print("Done.")
    print(f"Experiment signature: {signature}")
    print(f"Results saved to: {output_dir}")


if __name__ == "__main__":
    main()
